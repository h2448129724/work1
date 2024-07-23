# -*- coding: utf-8 -*-

import json
import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

def get_date():
    return time.strftime("%Y-%m-%d")

def read_json_file(file_name):
    if os.path.exists(file_name):
        with open(file_name, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return []

def write_json_file(file_name, data):
    with open(file_name, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def read_excel(file_name):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

class BrowserAutomation:
    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--no-sandbox")  # 禁用沙盒模式
        self.driver = webdriver.Chrome(options=options)

    def open_browser(self, url):
        self.driver.get(url)

    def login_gv(self, email, password):
        self.driver.get('https://accounts.google.com/')
        email_input = self.driver.find_element(By.ID, 'identifierId')
        email_input.send_keys(email)
        self.driver.find_element(By.ID, 'identifierNext').click()
        print("等待密码输入框出现...")
        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.NAME, 'Passwd')))
        password_input = self.driver.find_element(By.NAME, 'Passwd')
        password_input.send_keys(password)
        self.driver.find_element(By.ID, 'passwordNext').click()
        WebDriverWait(self.driver, 30).until(EC.title_contains('Google Account'))

    def send_message(self, phone, message):
        self.driver.get('https://voice.google.com/')
        print("等待消息按钮出现...")
        message_button = WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "Messages")]'))
        )
        message_button.click()
        print("等待新消息按钮出现...")
        new_message_button = WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@aria-label="Send new message"]'))
        )
        new_message_button.click()
        print("等待号码输入框出现...")
        phone_input = WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@aria-label="Enter name or number"]'))
        )
        phone_input.send_keys(phone)
        phone_input.send_keys(Keys.RETURN)
        print("等待消息输入框出现...")
        message_input = WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@aria-label="Send a message"]'))
        )
        message_input.send_keys(message)
        send_button = self.driver.find_element(By.XPATH, '//*[@aria-label="Send message"]')
        send_button.click()

    def close_browser(self):
        self.driver.quit()

if __name__ == "__main__":
    # 配置和初始化
    EMAIL = "your_email@gmail.com"
    PASSWORD = "your_password"
    PHONE_NUMBER = "recipient_phone_number"  # 目标电话号码
    MESSAGE_CONTENT = "Hello, this is a test message."  # 消息内容

    # 初始化服务
    browser_automation = BrowserAutomation()

    # 打开浏览器并登录GV
    browser_automation.open_browser('https://accounts.google.com/')
    browser_automation.login_gv(EMAIL, PASSWORD)
    # 发送消息
    browser_automation.send_message(PHONE_NUMBER, MESSAGE_CONTENT)
    # 关闭浏览器
    browser_automation.close_browser()
